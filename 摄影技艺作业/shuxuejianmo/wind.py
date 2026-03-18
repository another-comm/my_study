import cdsapi
import xarray as xr
import pandas as pd
import matplotlib.pyplot as plt
import os
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 下载 ERA5 风速数据
def download_era5_wind_month(year, month, save_path):
    print(f"📥 开始下载 {year}-{month:02d} 的 ERA5 风速数据...")
    try:
        c = cdsapi.Client()
        days_in_month = calendar.monthrange(year, month)[1]
        days = [str(day).zfill(2) for day in range(1, days_in_month + 1)]

        c.retrieve(
            'reanalysis-era5-single-levels',
            {
                'product_type': 'reanalysis',
                'variable': [
                    '10m_u_component_of_wind',
                    '10m_v_component_of_wind'
                ],
                'year': str(year),
                'month': str(month).zfill(2),
                'day': days,
                'time': [f"{str(h).zfill(2)}:00" for h in range(24)],
                'format': 'netcdf',
                'area': [32, 120, 31, 121],  # [北, 西, 南, 东]
                'grid': [0.25, 0.25],
            },
            save_path
        )
        print(f"✅ 下载完成：{save_path}")
    except Exception as e:
        print("❌ 下载失败：", e)
        raise

# 美化 Excel
def beautify_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = align_center
            cell.border = thin_border
        cell = row[1]
        if isinstance(cell.value, (int, float)):
            cell.value = f"{cell.value:.2f} m/s"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    wb.save(filepath)
    print(f"✨ Excel 文件已美化：{filepath}")

# 提取风速并导出 Excel
def extract_wind_to_excel(nc_file, excel_file):
    print(f"📖 正在读取：{nc_file}")
    try:
        ds = xr.open_dataset(nc_file, engine='netcdf4')

        u10 = ds['u10'].sel(latitude=31.5, longitude=120.5, method='nearest')
        v10 = ds['v10'].sel(latitude=31.5, longitude=120.5, method='nearest')

        wind_speed = (u10 ** 2 + v10 ** 2) ** 0.5

        df = wind_speed.to_series().reset_index()
        df.columns = ['时间', '10米风速（m/s）']

        df.to_excel(excel_file, index=False)
        beautify_excel(excel_file)
        return df
    except Exception as e:
        print("❌ 处理失败：", e)
        raise

# 绘图
def plot_wind(df):
    print("📊 开始绘图...")
    plt.figure(figsize=(14, 6))
    plt.plot(df['时间'], df['10米风速（m/s）'], color='green', linewidth=1)
    plt.title("无锡地区2022–2025年1–7月 10米风速变化曲线", fontsize=16)
    plt.xlabel("时间", fontsize=14)
    plt.ylabel("风速 (m/s)", fontsize=14)
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# 主函数
if __name__ == "__main__":
    years = range(2021, 2026)
    months = range(1, 8)
    base_path = r"C:\Users\32610\Desktop\shuxuejianmo"
    all_dfs = []

    print("📁 当前目录：", os.getcwd())
    print("🚀 正在批量获取 ERA5 风速数据...")

    for year in years:
        for month in months:
            nc_file = os.path.join(base_path, f"era5_wind_{year}_{month:02d}.nc")
            excel_file = os.path.join(base_path, f"wind_wuxi_{year}_{month:02d}.xlsx")

            try:
                if not os.path.exists(nc_file):
                    download_era5_wind_month(year, month, nc_file)
                else:
                    print(f"✅ 文件已存在：{nc_file}")
            except Exception:
                print(f"❌ 下载 {year}-{month:02d} 失败，跳过。")
                continue

            try:
                df = extract_wind_to_excel(nc_file, excel_file)
                all_dfs.append(df)
            except Exception:
                print(f"❌ 处理 {year}-{month:02d} 数据失败。")

    if all_dfs:
        full_df = pd.concat(all_dfs, ignore_index=True)
        plot_wind(full_df)
    else:
        print("⚠️ 无有效数据，未能绘图。")
