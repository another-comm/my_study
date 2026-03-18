import cdsapi
import xarray as xr
import pandas as pd
import matplotlib.pyplot as plt
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import calendar

# 下载 ERA5 指定年月的整月日照数据
def download_era5_ssrd_month(year, month, save_path):
    print(f"开始下载 {year}-{month:02d} 的 ERA5 日照数据...")
    try:
        c = cdsapi.Client()
        days_in_month = calendar.monthrange(year, month)[1]
        days = [str(day).zfill(2) for day in range(1, days_in_month + 1)]

        c.retrieve(
            'reanalysis-era5-single-levels',
            {
                'product_type': 'reanalysis',
                'variable': ['surface_solar_radiation_downwards'],
                'year': str(year),
                'month': str(month).zfill(2),
                'day': days,
                'time': [f"{str(h).zfill(2)}:00" for h in range(24)],
                'format': 'netcdf',
                'area': [32, 120, 31, 121],  # [北, 西, 南, 东]
                'grid': [0.25, 0.25],
            },
            save_path
        )
        print(f"✅ 下载完成，文件保存在：{save_path}")
    except Exception as e:
        print("❌ 下载过程出现错误：", e)
        raise

# 美化 Excel 样式
def beautify_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    # 设置字体样式和居中对齐
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 格式化标题行
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # 单元格样式设置
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # 为第二列加上单位后缀（保留两位小数）
        cell = row[1]
        if isinstance(cell.value, (int, float)):
            cell.value = f"{cell.value:.2f} W/m²"

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:B{ws.max_row}"

    wb.save(filepath)
    print(f"✨ Excel 文件已美化：{filepath}")

# 读取 NetCDF，提取无锡点数据，导出 Excel
def extract_ssrd_to_excel(nc_file, excel_file):
    print(f"开始读取 NetCDF 文件：{nc_file}")
    try:
        ds = xr.open_dataset(nc_file, engine='netcdf4')

        # 选取无锡对应的网格点：纬度31.5， 经度120.5（可根据网格调整）
        # 用最近邻插值选点
        ssrd = ds['ssrd'].sel(latitude=31.5, longitude=120.5, method='nearest')

        # ssrd 单位是 J/m² 逐小时累积，转换成 W/m²（J/s/m²，3600秒每小时）
        ssrd_wm2 = ssrd / 3600

        # 转为 pandas.Series 并重置索引，得到时间列
        df = ssrd_wm2.to_series().reset_index()
        df.columns = ['时间', '下行太阳辐射（W/m²）']

        # 导出 Excel
        df.to_excel(excel_file, index=False)
        beautify_excel(excel_file)

        return df
    except Exception as e:
        print("❌ 读取或转换 NetCDF 文件时出现错误：", e)
        raise

# 画图函数
def plot_ssrd(df):
    print("开始绘制图形...")
    plt.figure(figsize=(12, 6))
    plt.plot(df['时间'], df['下行太阳辐射（W/m²）'], marker='o', color='orange', linewidth=1)
    plt.title("无锡地区2025年1月至7月日照辐射变化曲线", fontsize=16)
    plt.xlabel("时间", fontsize=14)
    plt.ylabel("辐射强度 (W/m²)", fontsize=14)
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# 主函数
if __name__ == "__main__":
    year = 2025
    months = range(1, 8)  # 1月至7月
    base_path = r"C:\Users\32610\Desktop\shuxuejianmo"
    all_dfs = []

    print("📁 当前工作目录:", os.getcwd())
    print("🔧 开始批量下载和处理数据...")

    for month in months:
        nc_file = os.path.join(base_path, f"era5_wuxi_{year}_{month:02d}.nc")
        excel_file = os.path.join(base_path, f"ssrd_wuxi_{year}_{month:02d}.xlsx")

        # 下载
        try:
            if not os.path.exists(nc_file):
                download_era5_ssrd_month(year, month, nc_file)
            else:
                print(f"文件已存在，跳过下载：{nc_file}")
        except Exception:
            print(f"❌ 下载 {year}-{month:02d} 失败，跳过此月。")
            continue

        # 读取处理
        if os.path.exists(nc_file):
            try:
                df = extract_ssrd_to_excel(nc_file, excel_file)
                all_dfs.append(df)
            except Exception:
                print(f"❌ 处理 {year}-{month:02d} 数据时发生错误。")
        else:
            print(f"❌ 文件 {nc_file} 不存在，无法读取。")

    # 合并所有月份数据，绘制整体图形
    if all_dfs:
        full_df = pd.concat(all_dfs, ignore_index=True)
        plot_ssrd(full_df)
    else:
        print("❌ 没有可用数据绘图。")
