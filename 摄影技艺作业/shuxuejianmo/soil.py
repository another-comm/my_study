import cdsapi
import xarray as xr
import pandas as pd
import matplotlib.pyplot as plt
import os
import zipfile
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 下载 ERA5 土壤湿度（7-28cm）
def download_era5_soil_moisture_layer2(year, month, save_path):
    print(f"📥 下载 {year}-{month:02d} ERA5 土壤湿度（Layer 2）...")
    try:
        c = cdsapi.Client()
        days_in_month = calendar.monthrange(year, month)[1]
        days = [str(day).zfill(2) for day in range(1, days_in_month + 1)]

        zip_path = save_path.replace(".nc", ".zip")

        c.retrieve(
            'reanalysis-era5-land',
            {
                'product_type': 'reanalysis',
                'variable': ['volumetric_soil_water_layer_2'],
                'year': str(year),
                'month': str(month).zfill(2),
                'day': days,
                'time': [f"{str(h).zfill(2)}:00" for h in range(24)],
                'format': 'netcdf',
                'area': [32, 120, 31, 121],  # [北, 西, 南, 东]
                'grid': [0.25, 0.25],
            },
            zip_path
        )

        # 解压并重命名
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(os.path.dirname(save_path))
            for name in zip_ref.namelist():
                extracted_file = os.path.join(os.path.dirname(save_path), name)
                os.rename(extracted_file, save_path)
        os.remove(zip_path)

        print(f"✅ 解压并保存 NetCDF：{save_path}")
    except Exception as e:
        print("❌ 下载失败：", e)
        raise

# 美化 Excel 样式
def beautify_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        val = row[1].value
        if isinstance(val, (int, float)):
            row[1].value = f"{val:.4f} m³/m³"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    wb.save(filepath)
    print(f"✨ Excel 已美化：{filepath}")

# 读取土壤湿度 swvl2（中浅层）
def extract_soil_moisture_to_excel(nc_file, excel_file):
    print(f"📖 正在读取：{nc_file}")
    try:
        ds = xr.open_dataset(nc_file, engine='netcdf4')
        sm = ds['swvl2'].sel(latitude=31.5, longitude=120.5, method='nearest')

        df = sm.to_series().reset_index()
        df.columns = ['时间', '中浅层土壤湿度（m³/m³）']

        df.to_excel(excel_file, index=False)
        beautify_excel(excel_file)

        return df
    except Exception as e:
        print("❌ 读取失败：", e)
        raise

# 绘图
def plot_soil_moisture(df):
    print("📊 开始绘图...")
    plt.figure(figsize=(14, 6))
    plt.plot(df['时间'], df['中浅层土壤湿度（m³/m³）'], color='blue', linewidth=1)
    plt.title("无锡地区2021–2025年1–7月 中浅层土壤湿度变化", fontsize=16)
    plt.xlabel("时间", fontsize=14)
    plt.ylabel("土壤湿度 (m³/m³)", fontsize=14)
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# 主函数
if __name__ == "__main__":
    years = range(2021, 2026)
    months = range(6, 8)
    base_path = r"C:\Users\32610\Desktop\shuxuejianmo"
    all_dfs = []

    print("🚀 批量处理 ERA5 土壤湿度（中浅层 swvl2）数据...")

    for year in years:
        for month in months:
            nc_file = os.path.join(base_path, f"era5_soilmoisture2_{year}_{month:02d}.nc")
            excel_file = os.path.join(base_path, f"soilmoisture2_wuxi_{year}_{month:02d}.xlsx")

            try:
                if not os.path.exists(nc_file):
                    download_era5_soil_moisture_layer2(year, month, nc_file)
                else:
                    print(f"✅ 文件已存在：{nc_file}")
            except Exception:
                print(f"❌ 下载 {year}-{month:02d} 失败，跳过。")
                continue

            try:
                df = extract_soil_moisture_to_excel(nc_file, excel_file)
                all_dfs.append(df)
            except Exception:
                print(f"❌ 处理 {year}-{month:02d} 数据失败。")

    if all_dfs:
        full_df = pd.concat(all_dfs, ignore_index=True)
        plot_soil_moisture(full_df)
    else:
        print("⚠️ 没有可用数据，未能绘图。")

